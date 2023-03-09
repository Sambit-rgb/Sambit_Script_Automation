#########################################################################################################
# working code DONT TOUCH
import subprocess
import openpyxl
import xlrd
import requests
import urllib3
import time
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from xlutils.copy import copy


fil = "C:/Users/11014494/Desktop/tzt.xlsx"
workbook = xlrd.open_workbook(fil)
sheet = workbook.sheet_by_index(0)

fil_u = "C:/Users/11014494/Desktop/credential.xlsx"
workbook1 = xlrd.open_workbook(fil_u)
sheet1 = workbook1.sheet_by_index(0)

write = openpyxl.load_workbook("C:/Users/11014494/Desktop/credential.xlsx")
write1 = openpyxl.load_workbook("C:/Users/11014494/Desktop/tzt.xlsx")
write.create_sheet(title='Data')
sh = write['Data']
sh1 = write['Password']
sh2 = write1['Sheet1']
rows = sh.max_row
rows1 = sh1.max_row
rows2 = sh2.max_row
col = sh.max_column
col1 = sh1.max_column
col2 = sh2.max_column

list0 = []
list1 = []
list3 = []

for r in range(1, rows2+1):
    for w in range(1, col2+1):
        c = sh2.cell(r,w)
        url = c.value
        data = url.split("//")
        for o in range(1, rows1 + 1):
            d = sh1.cell(column=1, row=o)
            t = d.value
            n = sh1.cell(column=2, row=o)
            y = n.value


            x = data[0] + '//' + str(t) + ':' + str(y) + '@' + data[1]
            print (x)
            #list3.append(x)

            try:
                h = requests.get(x, allow_redirects=False, verify=False)
                z = h.headers
                code = h.status_code
                list0.append(z)
                list1.append(code)
                list3.append(x)

            except requests.exceptions.RequestException as e:
                z = t.headers
                code = t.status_code
                list0.append(z)
                list1.append(code)
                list3.append(x)


for i in range(1, sheet1.nrows * sheet.nrows + 1):
    cll1 = sh.cell(i, 2)
    url_header = list0[i - 1]
    cll1.value = str(url_header)
    #print (url_header)


for j in range(1, sheet1.nrows * sheet.nrows + 1):
    cll2 = sh.cell(j, 3)
    s_code = list1[j - 1]
    cll2.value = s_code
    #print (s_code)


for k in range(1, sheet1.nrows * sheet.nrows + 1):
    cll3 = sh.cell(k, 1)
    txt = list3[k - 1]

    cll3.value = txt




write.save("C:/Users/11014494/Desktop/credential.xlsx")







