import openpyxl
#working code DONT TOUCH
import xlrd
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from xlutils.copy import copy

fil = "C:/Users/11014494/Desktop/ttt.xlsx"
workbook = xlrd.open_workbook(fil)
wb = copy(workbook)
w_sheet = wb.get_sheet(0)
sheet = workbook.sheet_by_index(0)



write = openpyxl.load_workbook("C:/Users/11014494/Desktop/ttt.xlsx")
sh = write["Sheet1"]
rows = sh.max_row
list = []


for row in range(sheet.nrows):
    url = sheet.cell_value(row, 0)

    try:
       t = requests.get(url, allow_redirects=False, verify=False)

       z =t.status_code
       list.append(z)


    except requests.exceptions.RequestException as e:
        z = t.status_code
        list.append(z)

for i in range(1, rows + 1):
    cell = sh.cell(i, 2)
    x = list[i - 1]
    #print (x)
    cell.value = x
    write.save("C:/Users/11014494/Desktop/ttt.xlsx")



