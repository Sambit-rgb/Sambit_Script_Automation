#########################################################################################################
# working code DONT TOUCH
import subprocess
import openpyxl
import xlrd
import requests
import urllib3
import time
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from xlutils.copy import copy

fil = "C:/Users/11014494/Desktop/ttt.xlsx"
workbook = xlrd.open_workbook(fil)
sheet = workbook.sheet_by_index(0)

fil_u = "C:/Users/11014494/Desktop/credential.xlsx"
workbook1 = xlrd.open_workbook(fil_u)
sheet1 = workbook1.sheet_by_index(0)


write = openpyxl.load_workbook("C:/Users/11014494/Desktop/credential.xlsx")
write.create_sheet(title='Data')
sh = write['Data']
rows = sh.max_row

list = []
list1 = []
list3 = []

for row in range(sheet.nrows):
        url = sheet.cell_value(row, 0)
        data = url.split("//")
        for row in range(sheet1.nrows):
            t = sheet1.cell_value(row, 0)
            y = sheet1.cell_value(row, 1)
            x = data[0] + '//' + t + ':' + y + '@' + data[1]
            #list3.append(x)

            try:
                t = requests.get(x, allow_redirects=False, verify=False)
                z = t.headers
                code = t.status_code
                list.append(z)
                list1.append(code)
                list3.append(x)



            except requests.exceptions.RequestException as e:
                z = t.headers
                code = t.status_code
                list.append(z)
                list1.append(code)
                list3.append(x)


for i in range(1, sheet1.nrows * sheet.nrows + 1):
    cll1 = sh.cell(i, 2)
    url_header = list[i - 1]
    cll1.value = str(url_header)
    #print (url_header)


for j in range(1, sheet1.nrows * sheet.nrows + 1):
    cll2 = sh.cell(j, 3)
    s_code = list1[j - 1]
    cll2.value = s_code
    #print (s_code)


for k in range(1, sheet1.nrows * sheet.nrows + 1):
    cll3 = sh.cell(k, 1)
    txt = list3[k - 1]
    #print (txt)
    cll3.value = txt




write.save("C:/Users/11014494/Desktop/credential.xlsx")







